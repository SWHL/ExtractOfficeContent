# -*- encoding: utf-8 -*-
# @Author: SWHL
# @Contact: liekkaskono@163.com
from io import BytesIO
import argparse
import tempfile
import uuid
import warnings
import zipfile
from pathlib import Path
from typing import List, Union

import openpyxl
import pandas as pd
from openpyxl.workbook.workbook import Workbook

from .utils import mkdir


class ExtractExcel():
    def __init__(self,):
        self.img_suffix = [".jpg", ".jpeg", ".png", ".bmp"]

    def __call__(self, excel_content: Union[str, Path, bytes],
                 out_format: str = 'markdown',
                 save_img_dir: str = None) -> List:

        if isinstance(excel_content, (str, Path)):
            if not Path(excel_content).exists():
                raise FileNotFoundError(f'{excel_content} does not exist.')
            excel_content = str(excel_content)
        elif isinstance(excel_content, bytes):
            excel_content = BytesIO(excel_content)

        wb = self.unmerge_cell(excel_content)
        data_table = self.extract_table(wb, out_format)

        if save_img_dir:
            try:
                self.extract_imgs(excel_content, save_img_dir)
            except FileExistsError:
                warnings.warn(f'The {excel_content} does not contain any images.')

        return data_table

    def unmerge_cell(self, file_name: str) -> Workbook:
        wb = openpyxl.load_workbook(file_name)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            self.unmerge_and_fill_cells(sheet)
        return wb

    def unmerge_and_fill_cells(self, worksheet: Workbook) -> None:
        """
        # 拆分所有的合并单元格，并赋予合并之前的值。
        # 由于openpyxl并没有提供拆分并填充的方法，所以使用该方法进行完成
        """
        all_merged_cell_ranges = list(worksheet.merged_cells.ranges)

        for merged_cell_range in all_merged_cell_ranges:
            merged_cell = merged_cell_range.start_cell
            worksheet.unmerge_cells(range_string=merged_cell_range.coord)

            for row_index, col_index in merged_cell_range.cells:
                cell = worksheet.cell(row=row_index, column=col_index)
                cell.value = merged_cell.value

    def extract_table(self, wb: Workbook, out_format: str) -> List:
        sheet_names = wb.sheetnames
        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp_save_path = Path(tmp_dir) / f'{uuid.uuid1()}.xlsx'
            wb.save(str(tmp_save_path))
            wb.close()

            df_datas = []
            for name in sheet_names:
                data = pd.read_excel(tmp_save_path, index_col=None,
                                     sheet_name=name)
                cvt_data = self.convert_table(data, out_format)
                df_datas.append(cvt_data)
        return df_datas

    @staticmethod
    def convert_table(df_table: pd.core.frame.DataFrame,
                      out_format: str) -> str:
        if 'to_' not in out_format:
            out_format = f'to_{out_format}'

        try:
            return getattr(df_table, out_format)()
        except AttributeError as exc:
            raise AttributeError(f'{out_format} is not supported.') from exc

    def extract_imgs(self, excel_content: Union[str, Path, bytes],
                     save_img_dir: Union[str, Path]) -> None:
        with zipfile.ZipFile(excel_content) as zf:
            file_list = zf.namelist()

            img_list = [path for path in file_list
                        if Path(path).suffix in self.img_suffix]

            if not img_list:
                raise FileExistsError('The xl/media is not existed.')

            mkdir(save_img_dir)
            for img_path in img_list:
                save_path = Path(save_img_dir) / Path(img_path).name
                with open(save_path, 'wb') as f:
                    f.write(zf.read(img_path))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('excel_path', type=str)
    parser.add_argument('-f', '--output_format', type=str, default='markdown',
                        choices=['markdown', 'html', 'latex', 'string'])
    parser.add_argument('-o', '--save_img_dir', type=str, default=None)
    args = parser.parse_args()

    excel_extract = ExtractExcel()
    res = excel_extract(args.excel_path, out_format=args.output_format,
                        save_img_dir=args.save_img_dir)
    print(res)


if __name__ == '__main__':
    main()
