# -*- encoding: utf-8 -*-
# @Author: SWHL
# @Contact: liekkaskono@163.com
import shutil
import tempfile
import uuid
import zipfile
from pathlib import Path
from typing import List, Union

import openpyxl
import pandas as pd
from openpyxl.workbook.workbook import Workbook

from .utils import mkdir


class ExtractExcel():
    def __init__(self,):
        pass

    def __call__(self, excel_path: str,
                 out_format: str = 'markdown',
                 is_save_img: bool = False,
                 save_img_dir: str = None) -> List:
        wb = self.unmerge_cell(excel_path)
        data_table = self.extract_table(wb, out_format)

        if is_save_img:
            if save_img_dir is None:
                raise ValueError(
                    'When is_save_img is True, save_img_dir must be not None.')
            self.extract_imgs(excel_path, save_img_dir)
        return data_table

    def unmerge_cell(self, file_name: str) -> Workbook:
        wb = openpyxl.load_workbook(file_name)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            self.unmerge_and_fill_cells(sheet)
        return wb

    def unmerge_and_fill_cells(self, worksheet: Workbook) -> None:
        """
        # 拆分所有的合并单元格，并赋予合并之前的值。
        # 由于openpyxl并没有提供拆分并填充的方法，所以使用该方法进行完成
        """
        all_merged_cell_ranges = list(worksheet.merged_cells.ranges)

        for merged_cell_range in all_merged_cell_ranges:
            merged_cell = merged_cell_range.start_cell
            worksheet.unmerge_cells(range_string=merged_cell_range.coord)

            for row_index, col_index in merged_cell_range.cells:
                cell = worksheet.cell(row=row_index, column=col_index)
                cell.value = merged_cell.value

    def extract_table(self, wb: Workbook, out_format: str) -> List:
        sheet_names = wb.sheetnames
        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp_save_path = Path(tmp_dir) / f'{uuid.uuid1()}.xlsx'
            wb.save(str(tmp_save_path))
            wb.close()

            df_datas = []
            for name in sheet_names:
                data = pd.read_excel(tmp_save_path, index_col=None,
                                     sheet_name=name)
                cvt_data = self.convert_table(data, out_format)
                df_datas.append(cvt_data)
        return df_datas

    @staticmethod
    def convert_table(df_table: pd.core.frame.DataFrame,
                      out_format: str) -> str:
        if 'to_' not in out_format:
            out_format = f'to_{out_format}'

        try:
            return getattr(df_table, out_format)()
        except AttributeError as exc:
            raise AttributeError(f'{out_format} is not supported.') from exc

    @staticmethod
    def extract_imgs(excel_path: str, save_img_dir: Union[str, Path]) -> None:
        excel_path = Path(excel_path)
        zip_excel_path = excel_path.with_name(f'{excel_path.stem}.zip')

        excel_path.rename(zip_excel_path)

        unzip_dir = excel_path.parent / excel_path.stem
        with zipfile.ZipFile(zip_excel_path) as zf:
            zf.extractall(unzip_dir)
        zip_excel_path.rename(excel_path)

        imgs_dir = unzip_dir / 'xl' / 'media'
        if not imgs_dir.exists():
            raise FileExistsError('The xl/media is not existed.')

        mkdir(save_img_dir)
        shutil.move(imgs_dir, save_img_dir)
        shutil.rmtree(unzip_dir)
