# -*- encoding: utf-8 -*-
# @Author: SWHL
# @Contact: liekkaskono@163.com
"""
测试pandas读取带有合并单元格的excel

"""
import openpyxl


def unmerge_and_fill_cells(worksheet):
    # 拆分所有的合并单元格，并赋予合并之前的值。
    # 由于openpyxl并没有提供拆分并填充的方法，所以使用该方法进行完成
    all_merged_cell_ranges = list(
        worksheet.merged_cells.ranges
    )

    for merged_cell_range in all_merged_cell_ranges:
        merged_cell = merged_cell_range.start_cell
        worksheet.unmerge_cells(range_string=merged_cell_range.coord)

        for row_index, col_index in merged_cell_range.cells:
            cell = worksheet.cell(row=row_index, column=col_index)
            cell.value = merged_cell.value


def unmerge_cell(filename):
    # 读取原始xlsx文件，拆分并填充单元格，然后生成中间临时文件。
    wb = openpyxl.load_workbook(filename)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        unmerge_and_fill_cells(sheet)

    filename = filename.replace(".xls", "_temp.xls")
    wb.save(filename)
    wb.close()

    # # openpyxl保存之后，再用pandas读取会存在公式无法读取到的情况，使用下面方式就可以了
    # # 如果你的excel不涉及公式，可以删除下面内容
    # # 原理为：使用windows打开excel，然后另存为一下
    # from win32com.client import Dispatch
    # xlApp = Dispatch("Excel.Application")
    # xlApp.Visible = False
    # xlBook = xlApp.Workbooks.Open(
    #     str(Path(".").absolute() / filename))  # 这里必须填绝对路径
    # xlBook.Save()
    # xlBook.Close()

    return filename


if __name__ == '__main__':
    excel_path = 'tests/test_files/excel_example.xlsx'
    unmerge_cell(excel_path)

    # data = pd.read_excel(excel_path, index_col=None)

    # print('ok')
